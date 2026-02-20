import os
import imageio_ffmpeg

# ✅ FORCE FFMPEG FOR WHISPER (MUST BE FIRST)
FFMPEG_EXE = imageio_ffmpeg.get_ffmpeg_exe()
os.environ["PATH"] = os.path.dirname(FFMPEG_EXE) + os.pathsep + os.environ.get("PATH", "")
os.environ["FFMPEG_BINARY"] = FFMPEG_EXE
if os.getenv("FLASK_DEBUG", "0") == "1":
    print("FFMPEG ready:", os.path.exists(FFMPEG_EXE))


from flask import Flask,Blueprint, render_template, request, redirect, url_for, session, flash, send_from_directory,current_app, jsonify,abort
from dotenv import load_dotenv
load_dotenv()
from flask_mysqldb import MySQL
from flask_bcrypt import Bcrypt
from datetime import datetime, timedelta
from pathlib import Path
#import datetime, json, io
import os
import csv
import shutil
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import json
import zipfile
from werkzeug.utils import secure_filename
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import ctypes.util
from dysarthria_inference import predict_dysarthria
from flask_dance.contrib.google import make_google_blueprint, google


libc_name = ctypes.util.find_library("c")
if libc_name is None:
    libc_name = "msvcrt.dll"
import ctypes
ctypes.CDLL(libc_name)  # This call ensures the fallback library is used

import re
import uuid
import tempfile
import asyncio
import edge_tts
import whisper
# ✅ Force Whisper to use the bundled ffmpeg (instead of plain "ffmpeg")
# ✅ Force Whisper to decode audio using our bundled ffmpeg exe (fixes WinError 2)
try:
    import whisper.audio as wa
    import subprocess
    import numpy as np

    _orig_load_audio = wa.load_audio

    def load_audio_patched(file, sr=wa.SAMPLE_RATE):
        if isinstance(file, str):
            cmd = [
                FFMPEG_EXE,
                "-nostdin", "-threads", "0",
                "-i", file,
                "-f", "s16le",
                "-ac", "1",
                "-acodec", "pcm_s16le",
                "-ar", str(sr),
                "-"
            ]
            out = subprocess.run(cmd, capture_output=True, check=True).stdout
            audio = np.frombuffer(out, np.int16).flatten().astype(np.float32) / 32768.0
            return audio
        return _orig_load_audio(file, sr=sr)

    wa.load_audio = load_audio_patched
    print("✅ Whisper load_audio patched to use:", FFMPEG_EXE)
except Exception as e:
    print("⚠️ Whisper load_audio patch failed:", e)


import librosa
import numpy as np
import nltk
from textblob import TextBlob
from functools import partial


# ✅ alias the module so it won't shadow the class import elsewhere

import json, io

import soundfile as sf

from audio_enhance import autotune_enhance_audio


import pyttsx3

_tts_engine = None

def tts_pyttsx3_wav(text: str, out_wav: str):
    global _tts_engine
    if _tts_engine is None:
        _tts_engine = pyttsx3.init()
        _tts_engine.setProperty("rate", 170)  # adjust speed if needed
    _tts_engine.save_to_file(text, out_wav)
    _tts_engine.runAndWait()




# from pydrive.auth import GoogleAuth
# from pydrive.drive import GoogleDrive

app = Flask(__name__)
import os
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
# Google OAuth Config
app.config['OAUTHLIB_INSECURE_TRANSPORT'] = True  # Only for development

GOOGLE_OAUTH_CLIENT_ID = os.getenv("GOOGLE_OAUTH_CLIENT_ID")
GOOGLE_OAUTH_CLIENT_SECRET = os.getenv("GOOGLE_OAUTH_CLIENT_SECRET")

google_bp = make_google_blueprint(
    client_id=GOOGLE_OAUTH_CLIENT_ID,
    client_secret=GOOGLE_OAUTH_CLIENT_SECRET,
    scope=[
        "openid",
        "https://www.googleapis.com/auth/userinfo.profile",
        "https://www.googleapis.com/auth/userinfo.email"
    ],
    redirect_to="google_login"
)
# Set the authorization parameter to force a fresh account selection.
google_bp.session.authorization_url_params = {"prompt": "select_account"}

app.register_blueprint(google_bp, url_prefix="/login")

# app.register_blueprint(google_bp, url_prefix="/login")
app.secret_key = os.environ['SECRET_KEY']
# MySQL Configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = 'ABDraf12@'
app.config['MYSQL_DB'] = 'login'
app.config['ARCHIVE_FOLDER'] = 'archives'
app.config['UPLOAD_DIR'] = 'audio_uploads'

app.config['UPLOAD_FOLDER'] = 'uploads'

# UPLOAD_DIR = 'audio_uploads'
os.makedirs(app.config['UPLOAD_DIR'], exist_ok=True)


mysql = MySQL(app)
bcrypt = Bcrypt(app)

@app.route("/api/patients", methods=["GET"])
def api_patients():
    """
    Returns real patients from DB for therapist dashboard dropdown.
    """
    try:
        cur = mysql.connection.cursor()

        # ✅ If you have a role column and patients are stored as 'patient'
        cur.execute("""
            SELECT id, username, email
            FROM users
            WHERE role = 'patient'
            ORDER BY username ASC
        """)

        rows = cur.fetchall()
        cur.close()

        patients = []
        for r in rows:
            patients.append({
                "id": r[0],
                "username": r[1],
                "email": r[2]
            })

        return jsonify({"success": True, "patients": patients}), 200

    except Exception as e:
        print("api_patients error:", e)
        return jsonify({"success": False, "error": "Failed to fetch patients"}), 500


def get_user_id(username):
    cur = mysql.connection.cursor()
    cur.execute("SELECT id FROM users WHERE username = %s", (username,))
    row = cur.fetchone()
    cur.close()
    return row[0] if row else None

def get_or_create_users_id(identifier):
    """
    Returns users.id for logged-in user.
    identifier may be username OR email.
    Creates users row if missing.
    """
    if not identifier:
        return None

    cur = mysql.connection.cursor()

    # 1) Try users table
    cur.execute(
        "SELECT id FROM users WHERE username = %s OR email = %s",
        (identifier, identifier)
    )
    row = cur.fetchone()
    if row:
        cur.close()
        return row[0]

    # 2) Try login table
    cur.execute(
        "SELECT username, email, password FROM login WHERE username = %s OR email = %s",
        (identifier, identifier)
    )
    login_row = cur.fetchone()

    if not login_row:
        cur.close()
        return None

    username, email, password = login_row

    # 3) Insert into users
    try:
        cur.execute("""
            INSERT INTO users (email, password, username)
            VALUES (%s, %s, %s)
        """, (email, password, username))

        mysql.connection.commit()
        new_id = cur.lastrowid
        cur.close()
        return new_id

    except Exception:
        mysql.connection.rollback()
        cur.execute("SELECT id FROM users WHERE email = %s", (email,))
        row = cur.fetchone()
        cur.close()
        return row[0] if row else None


def db_get_uploaded_audio_by_id(mysql, audio_id: int, user_id: int):
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id, user_id, original_filename, stored_filename, file_path
        FROM uploaded_audio
        WHERE id=%s AND user_id=%s
        LIMIT 1
    """, (audio_id, user_id))
    row = cur.fetchone()
    cur.close()
    return row

def db_upsert_stutter_removed(mysql, user_id: int, uploaded_audio_id: int,
                             original_filename: str, cleaned_filename: str,
                             cleaned_audio_bytes: bytes, raw_text: str, clean_text: str):
    cur = mysql.connection.cursor()

    # keep 1 result per (user_id, uploaded_audio_id)
    cur.execute("""
        SELECT id FROM stutter_removed_audio
        WHERE user_id=%s AND uploaded_audio_id=%s
        LIMIT 1
    """, (user_id, uploaded_audio_id))
    existing = cur.fetchone()

    if existing:
        cur.execute("""
            UPDATE stutter_removed_audio
            SET original_filename=%s,
                cleaned_filename=%s,
                cleaned_audio=%s,
                raw_text=%s,
                clean_text=%s,
                created_at=NOW()
            WHERE id=%s
        """, (original_filename, cleaned_filename, cleaned_audio_bytes, raw_text, clean_text, existing[0]))
        result_id = existing[0]
    else:
        cur.execute("""
            INSERT INTO stutter_removed_audio
            (user_id, uploaded_audio_id, original_filename, cleaned_filename, cleaned_audio, raw_text, clean_text, created_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s,NOW())
        """, (user_id, uploaded_audio_id, original_filename, cleaned_filename, cleaned_audio_bytes, raw_text, clean_text))
        result_id = cur.lastrowid

    mysql.connection.commit()
    cur.close()
    return result_id


def _is_therapist_logged_in():
    identifier = session.get("username")  # you store username/email here
    if not identifier:
        return False

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT role FROM users
        WHERE username=%s OR email=%s
        LIMIT 1
    """, (identifier, identifier))
    row = cur.fetchone()
    cur.close()

    return bool(row and (row[0] or "").lower() == "therapist")


def _is_patient_logged_in():
    identifier = session.get("username")
    if not identifier:
        return False

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT role FROM users
        WHERE username=%s OR email=%s
        LIMIT 1
    """, (identifier, identifier))
    row = cur.fetchone()
    cur.close()

    return bool(row and (row[0] or "").lower() == "patient")

@app.route("/therapist/notes/save", methods=["POST"])
def save_therapist_note():
    try:
        # ✅ Only therapists can save notes
        if not _is_therapist_logged_in():
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        identifier = session.get("username")  # can be username or email (google)
        if not identifier:
            return jsonify({"success": False, "error": "Not logged in"}), 401

        # ✅ get therapist_id from users table
        cur = mysql.connection.cursor()
        cur.execute("""
            SELECT id
            FROM users
            WHERE username=%s OR email=%s
            LIMIT 1
        """, (identifier, identifier))
        trow = cur.fetchone()
        if not trow:
            cur.close()
            return jsonify({"success": False, "error": "Therapist not found"}), 404

        therapist_id = trow[0]

        # ✅ get form values
        patient_id = (request.form.get("patient_id") or "").strip()
        subject    = (request.form.get("subject") or "").strip()
        message    = (request.form.get("message") or "").strip()

        if not patient_id or not subject or not message:
            cur.close()
            return jsonify({"success": False, "error": "All fields are required"}), 400

        # ✅ verify selected user is a patient
        cur.execute("SELECT role FROM users WHERE id=%s LIMIT 1", (patient_id,))
        prow = cur.fetchone()
        if not prow or (prow[0] or "").lower() != "patient":
            cur.close()
            return jsonify({"success": False, "error": "Selected user is not a patient"}), 400

        # ✅ insert note
        cur.execute("""
            INSERT INTO therapist_notes (therapist_id, patient_id, subject, message, created_at)
            VALUES (%s, %s, %s, %s, NOW())
        """, (therapist_id, patient_id, subject, message))
        mysql.connection.commit()
        cur.close()

        return jsonify({"success": True}), 200

    except Exception as e:
        print("save_therapist_note error:", e)
        return jsonify({"success": False, "error": "Failed to save note"}), 500


@app.route("/therapist/notes/<int:note_id>/delete", methods=["POST"])
def delete_therapist_note(note_id):
    try:
        if not _is_therapist_logged_in():
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        identifier = session.get("username")
        if not identifier:
            return jsonify({"success": False, "error": "Not logged in"}), 401

        cur = mysql.connection.cursor()

        # get therapist_id
        cur.execute("""
            SELECT id FROM users
            WHERE username=%s OR email=%s
            LIMIT 1
        """, (identifier, identifier))
        trow = cur.fetchone()
        therapist_id = trow[0] if trow else None
        if not therapist_id:
            cur.close()
            return jsonify({"success": False, "error": "Therapist not found"}), 404

        # delete only if this note belongs to this therapist
        cur.execute("""
            DELETE FROM therapist_notes
            WHERE id=%s AND therapist_id=%s
        """, (note_id, therapist_id))
        mysql.connection.commit()

        deleted = cur.rowcount
        cur.close()

        if deleted == 0:
            return jsonify({"success": False, "error": "Note not found"}), 404

        return jsonify({"success": True}), 200

    except Exception as e:
        print("delete_therapist_note error:", e)
        return jsonify({"success": False, "error": "Delete failed"}), 500


@app.route("/therapist/speech/assign", methods=["POST"])
def assign_speech_exercise():
    try:
        # ✅ Therapist only
        if not _is_therapist_logged_in():
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        identifier = session.get("username")
        if not identifier:
            return jsonify({"success": False, "error": "Not logged in"}), 401

        patient_id  = request.form.get("patient_id")
        exercise_id = request.form.get("exercise_id")
        due_date    = request.form.get("due_date") or None

        if not patient_id or not exercise_id:
            return jsonify({"success": False, "error": "Missing fields"}), 400

        cur = mysql.connection.cursor()

        # ✅ Get therapist_id
        cur.execute("""
            SELECT id FROM users
            WHERE username=%s OR email=%s
            LIMIT 1
        """, (identifier, identifier))
        trow = cur.fetchone()
        if not trow:
            cur.close()
            return jsonify({"success": False, "error": "Therapist not found"}), 404

        therapist_id = trow[0]

        # ✅ Verify patient
        cur.execute("SELECT role FROM users WHERE id=%s LIMIT 1", (patient_id,))
        prow = cur.fetchone()
        if not prow or prow[0].lower() != "patient":
            cur.close()
            return jsonify({"success": False, "error": "Invalid patient"}), 400

        # ✅ Insert assignment
        cur.execute("""
            INSERT INTO speech_assignments
            (therapist_id, patient_id, exercise_id, due_date, status, assigned_at)
            VALUES (%s, %s, %s, %s, 'assigned', NOW())
        """, (therapist_id, patient_id, exercise_id, due_date))

        mysql.connection.commit()
        cur.close()

        return jsonify({"success": True}), 200

    except Exception as e:
        print("assign_speech_exercise error:", e)
        return jsonify({"success": False, "error": "Server error"}), 500

try:
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
except Exception as e:
    print(f"Failed to create upload directory: {e}")

try:
    os.makedirs(app.config['ARCHIVE_FOLDER'], exist_ok=True)
except Exception as e:
    print(f"Failed to create ARCHIVE_FOLDER directory: {e}")


#paient record blueprints:

# app = Flask(__name__)
# app.register_blueprint(reports_bp)
# app.register_blueprint(medication_bp)

@app.route('/predict_dysarthria', methods=['POST'])
def predict_dysarthria_route():
    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    import tempfile, uuid, os
    from werkzeug.utils import secure_filename

    try:
        with tempfile.TemporaryDirectory() as td:
            # Save uploaded file
            in_name = secure_filename(file.filename)
            in_path = os.path.join(td, in_name)
            file.save(in_path)

            # Convert to wav 16k mono (important for webm/mp3/m4a)
            wav_path = os.path.join(td, f"in_{uuid.uuid4().hex}.wav")
            to_wav16k_mono(in_path, wav_path)   # you already have this function

            # Predict
            label = predict_dysarthria(wav_path)

            return jsonify({"prediction": label}), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"{type(e).__name__}: {str(e)}"}), 500



# Google Drive Authentication
# gauth = GoogleAuth()
# gauth.LocalWebserverAuth()
# drive = GoogleDrive(gauth)


@app.route('/google_login')
def google_login():
    if not google.authorized:
        return redirect(url_for("google.login"))
    resp = google.get("/oauth2/v2/userinfo")
    user_info = resp.json()
    
    session['username'] = user_info["email"]
    flash('Logged in with Google!')
    return redirect(url_for('home'))


# Feature Title: Force Fresh Google Login
# Scenario Title: Create a custom route that forces the account chooser

from flask import redirect
from oauthlib.oauth2 import WebApplicationClient
# Feature Title: Force Fresh Google Login
# Scenario Title: Create a custom route that forces the account chooser even if the user is already logged in
from uuid import uuid4

# Feature Title: Force Fresh Google Login with Custom Authorization URL
# Scenario Title: Build a custom authorization URL that includes redirect_uri and forces account selection
@app.route('/force_google_login')
def force_google_login():
    # Clear any stored Google OAuth token and user data
    session.pop('google_oauth_token', None)
    session.pop('username', None)
    
    # Generate a new state value and store it for later verification
    from uuid import uuid4
    state = uuid4().hex
    session[google_bp.name + "_state"] = state

    # Get the underlying OAuth2 client from the blueprint's session
    client = google_bp.session._client

    # Build the external redirect URI using the correct callback endpoint.
    redirect_uri = url_for("google.authorized", _external=True)
    print("Redirect URI:", redirect_uri)  # This should print "http://localhost:5000/login/google/authorized"

    # Use the client's prepare_request_uri to manually build the authorization URL.
    auth_url = client.prepare_request_uri(
        "https://accounts.google.com/o/oauth2/auth",
        redirect_uri=redirect_uri,
        scope=google_bp.scope,
        state=state,
        prompt="select_account",   # Forces Google to prompt for account selection
        access_type="offline"       # Optional: requests a refresh token
    )
    
    return redirect(auth_url)


# File Upload Route
@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    username = session['username']
    user_dir = os.path.join(app.config['UPLOAD_FOLDER'], username)
    os.makedirs(user_dir, exist_ok=True)
    
    if request.method == 'POST':
        file = request.files['file']
        if file:
            file_path = os.path.join(user_dir, file.filename)
            file.save(file_path)
            flash('File uploaded successfully!')
    
    return render_template('upload.html')




@app.route('/profile', methods=['GET', 'POST'])
def profile():
    if 'username' not in session:
        flash('Please log in to access your profile.')
        return redirect(url_for('login'))

    identifier = session['username']  # could be username OR email (google login)

    cur = mysql.connection.cursor()

    # ✅ Fetch current user (include role)
    cur.execute("""
        SELECT id, username, full_name, password, dob, email, phone, address, bio, role
        FROM users
        WHERE username = %s OR email = %s
        LIMIT 1
    """, (identifier, identifier))
    user = cur.fetchone()

    if not user:
        cur.close()
        flash("User not found in users table.")
        return redirect(url_for("home"))

    user_id = user[0]

    if request.method == 'POST':
        username  = request.form.get('username')
        full_name = request.form.get('full_name')
        dob       = request.form.get('dob') or None
        email     = request.form.get('email')
        phone     = request.form.get('phone')
        address   = request.form.get('address')
        bio       = request.form.get('bio')
        role      = request.form.get('role', 'patient')  # ✅ NEW

        password  = request.form.get('password')  # optional
        password_db_value = None

        # ✅ If user typed a new password, hash it
        if password and password.strip():
            password_db_value = bcrypt.generate_password_hash(password).decode("utf-8")

        # ✅ UPDATE (not INSERT)
        if password_db_value:
            cur.execute("""
                UPDATE users
                SET username=%s, full_name=%s, dob=%s, email=%s,
                    phone=%s, address=%s, bio=%s, role=%s, password=%s
                WHERE id=%s
            """, (username, full_name, dob, email, phone, address, bio, role, password_db_value, user_id))
        else:
            cur.execute("""
                UPDATE users
                SET username=%s, full_name=%s, dob=%s, email=%s,
                    phone=%s, address=%s, bio=%s, role=%s
                WHERE id=%s
            """, (username, full_name, dob, email, phone, address, bio, role, user_id))

        mysql.connection.commit()
        cur.close()

        # ✅ keep session in sync if username changed
        session['username'] = username

        flash('Profile updated successfully!')
        return redirect(url_for('profile'))

    cur.close()

    # ✅ Convert tuple to dict so your Jinja works: user.username, user.role etc.
    user_dict = {
        "id": user[0],
        "username": user[1],
        "full_name": user[2],
        "password": user[3],
        "dob": user[4],
        "email": user[5],
        "phone": user[6],
        "address": user[7],
        "bio": user[8],
        "role": user[9],
    }

    return render_template('profile.html', user=user_dict)


@app.route("/api/patient/<int:patient_id>/uploaded_audio", methods=["GET"])
def api_patient_uploaded_audio(patient_id):
    # ✅ basic protection (therapist only)
    if not _is_therapist_logged_in():
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id, original_filename, stored_filename, uploaded_at, status
        FROM uploaded_audio
        WHERE user_id = %s
        ORDER BY uploaded_at DESC
    """, (patient_id,))
    rows = cur.fetchall()
    cur.close()

    audios = []
    for r in rows:
        audios.append({
            "id": r[0],
            "original_filename": r[1],
            "stored_filename": r[2],
            "uploaded_at": r[3].strftime("%Y-%m-%d %H:%M:%S") if r[3] else None,
            "status": r[4],
            # we will use a therapist-friendly stream route:
            "listen_url": url_for("therapist_stream_uploaded_audio", audio_id=r[0]),
        })

    return jsonify({"success": True, "audios": audios}), 200


@app.route("/therapist/audio/<int:audio_id>")
def therapist_stream_uploaded_audio(audio_id):
    if not _is_therapist_logged_in():
        abort(401)

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT file_path, stored_filename
        FROM uploaded_audio
        WHERE id=%s
        LIMIT 1
    """, (audio_id,))
    row = cur.fetchone()
    cur.close()

    if not row:
        abort(404)

    file_path, stored_filename = row
    return send_file(file_path, as_attachment=False, download_name=stored_filename)

@app.route("/dashboard")
def dashboard_redirect():
    if "username" not in session:
        flash("Please login first.")
        return redirect(url_for("login"))

    identifier = session.get("username")  # username OR email (google login)

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT role
        FROM users
        WHERE username = %s OR email = %s
        LIMIT 1
    """, (identifier, identifier))
    row = cur.fetchone()
    cur.close()

    # default fallback
    role = (row[0] if row and row[0] else "patient").lower()

    if role == "therapist":
        return redirect(url_for("therapist_dashboard"))
    else:
        return redirect(url_for("ui"))

@app.route("/api/speech/exercises", methods=["GET"])
def api_speech_exercises():
    if not _is_therapist_logged_in():
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    try:
        cur = mysql.connection.cursor()
        cur.execute("""
            SELECT id, title
            FROM speech_exercises
            WHERE is_active = 1
            ORDER BY title ASC
        """)
        rows = cur.fetchall()
        cur.close()

        exercises = [{"id": r[0], "title": r[1]} for r in rows]
        return jsonify({"success": True, "exercises": exercises}), 200

    except Exception as e:
        print("api_speech_exercises error:", e)
        return jsonify({"success": False, "error": "Failed to fetch exercises"}), 500

@app.route("/api/speech/assign", methods=["POST"])
def api_speech_assign():
    if not _is_therapist_logged_in():
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    identifier = session.get("username")
    if not identifier:
        return jsonify({"success": False, "error": "Not logged in"}), 401

    patient_id  = (request.form.get("patient_id") or "").strip()
    exercise_id = (request.form.get("exercise_id") or "").strip()
    due_date    = (request.form.get("due_date") or "").strip()  # can be ""

    if not patient_id or not exercise_id:
        return jsonify({"success": False, "error": "Patient and exercise are required"}), 400

    try:
        cur = mysql.connection.cursor()

        # therapist_id
        cur.execute("""
            SELECT id FROM users
            WHERE username=%s OR email=%s
            LIMIT 1
        """, (identifier, identifier))
        trow = cur.fetchone()
        if not trow:
            cur.close()
            return jsonify({"success": False, "error": "Therapist not found"}), 404
        therapist_id = trow[0]

        # verify patient role
        cur.execute("SELECT role FROM users WHERE id=%s LIMIT 1", (patient_id,))
        prow = cur.fetchone()
        if not prow or (prow[0] or "").lower() != "patient":
            cur.close()
            return jsonify({"success": False, "error": "Selected user is not a patient"}), 400

        # verify exercise exists
        cur.execute("SELECT id FROM speech_exercises WHERE id=%s AND is_active=1 LIMIT 1", (exercise_id,))
        ex = cur.fetchone()
        if not ex:
            cur.close()
            return jsonify({"success": False, "error": "Exercise not found"}), 404

        # insert assignment
        cur.execute("""
            INSERT INTO speech_assignments
            (therapist_id, patient_id, exercise_id, due_date, status, assigned_at)
            VALUES (%s, %s, %s, %s, 'assigned', NOW())
        """, (therapist_id, patient_id, exercise_id, (due_date or None)))

        mysql.connection.commit()
        new_id = cur.lastrowid
        cur.close()

        return jsonify({"success": True, "assignment_id": new_id}), 200

    except Exception as e:
        print("api_speech_assign error:", e)
        return jsonify({"success": False, "error": "Failed to assign exercise"}), 500

@app.route("/api/speech/submit/<int:assignment_id>", methods=["POST"])
def api_speech_submit(assignment_id):
    # ✅ patient only
    if not _is_patient_logged_in():
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    identifier = session.get("username")
    patient_id = get_or_create_users_id(identifier)
    if not patient_id:
        return jsonify({"success": False, "error": "User not found"}), 404

    if "file" not in request.files:
        return jsonify({"success": False, "error": "No file"}), 400

    f = request.files["file"]
    if not f or f.filename == "":
        return jsonify({"success": False, "error": "Empty filename"}), 400

    try:
        cur = mysql.connection.cursor()

        # ✅ verify assignment belongs to this patient
        cur.execute("""
            SELECT id
            FROM speech_assignments
            WHERE id=%s AND patient_id=%s
            LIMIT 1
        """, (assignment_id, patient_id))
        arow = cur.fetchone()
        if not arow:
            cur.close()
            return jsonify({"success": False, "error": "Assignment not found"}), 404

        # ✅ save file same way you save uploaded_audio
        original_filename = f.filename
        safe_name = secure_filename(original_filename)

        username_folder = session.get("username")  # ok
        user_dir = os.path.join(app.config["UPLOAD_DIR"], str(username_folder))
        os.makedirs(user_dir, exist_ok=True)

        stored_filename = f"{uuid.uuid4().hex}_{safe_name}"
        file_path = os.path.join(user_dir, stored_filename)

        f.save(file_path)

        # ✅ insert into uploaded_audio (reuse your existing table)
        cur.execute("""
            INSERT INTO uploaded_audio (user_id, original_filename, stored_filename, file_path, status, uploaded_at)
            VALUES (%s, %s, %s, %s, 'uploaded', NOW())
        """, (patient_id, original_filename, stored_filename, file_path))
        mysql.connection.commit()
        uploaded_audio_id = cur.lastrowid

        # ✅ link submission to assignment
        cur.execute("""
            INSERT INTO speech_submissions (assignment_id, patient_id, uploaded_audio_id, submitted_at, status)
            VALUES (%s, %s, %s, NOW(), 'submitted')
        """, (assignment_id, patient_id, uploaded_audio_id))
        mysql.connection.commit()

        # ✅ optionally mark assignment as submitted
        cur.execute("""
            UPDATE speech_assignments
            SET status='submitted'
            WHERE id=%s
        """, (assignment_id,))
        mysql.connection.commit()

        cur.close()
        return jsonify({"success": True}), 200

    except Exception as e:
        print("api_speech_submit error:", e)
        return jsonify({"success": False, "error": "Upload failed"}), 500

@app.route("/api/speech/submissions/<int:patient_id>", methods=["GET"])
def api_speech_submissions(patient_id):
    if not _is_therapist_logged_in():
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    identifier = session.get("username")
    if not identifier:
        return jsonify({"success": False, "error": "Not logged in"}), 401

    cur = mysql.connection.cursor()

    # therapist id
    cur.execute("""
        SELECT id FROM users
        WHERE username=%s OR email=%s
        LIMIT 1
    """, (identifier, identifier))
    trow = cur.fetchone()
    if not trow:
        cur.close()
        return jsonify({"success": False, "error": "Therapist not found"}), 404

    therapist_id = trow[0]

    # submissions for this therapist + this patient
    cur.execute("""
        SELECT
            ss.id AS submission_id,
            sa.id AS assignment_id,
            se.title AS exercise_title,
            ss.status,
            ss.submitted_at,
            ua.id AS uploaded_audio_id,
            ua.original_filename,
            ua.uploaded_at
        FROM speech_submissions ss
        JOIN speech_assignments sa ON sa.id = ss.assignment_id
        LEFT JOIN speech_exercises se ON se.id = sa.exercise_id
        JOIN uploaded_audio ua ON ua.id = ss.uploaded_audio_id
        WHERE sa.therapist_id = %s
          AND ss.patient_id = %s
        ORDER BY ss.submitted_at DESC
    """, (therapist_id, patient_id))

    rows = cur.fetchall()
    cur.close()

    subs = []
    for r in rows:
        subs.append({
            "submission_id": r[0],
            "assignment_id": r[1],
            "exercise_title": r[2] or "",
            "status": r[3] or "",
            "submitted_at": r[4].strftime("%Y-%m-%d %H:%M:%S") if r[4] else "",
            "uploaded_audio_id": r[5],
            "original_filename": r[6] or "audio",
            "uploaded_at": r[7].strftime("%Y-%m-%d %H:%M:%S") if r[7] else "",
            "listen_url": url_for("therapist_stream_uploaded_audio", audio_id=r[5]),
        })

    return jsonify({"success": True, "submissions": subs}), 200


@app.route("/api/speech/assignments/recent", methods=["GET"])
def api_speech_recent_assignments():
    if not _is_therapist_logged_in():
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    identifier = session.get("username")
    if not identifier:
        return jsonify({"success": False, "error": "Not logged in"}), 401

    try:
        cur = mysql.connection.cursor()

        # therapist_id
        cur.execute("""
            SELECT id FROM users
            WHERE username=%s OR email=%s
            LIMIT 1
        """, (identifier, identifier))
        trow = cur.fetchone()
        if not trow:
            cur.close()
            return jsonify({"success": False, "error": "Therapist not found"}), 404
        therapist_id = trow[0]

        cur.execute("""
            SELECT sa.id,
                   sa.patient_id,
                   u.username AS patient_username,
                   se.title AS exercise_title,
                   sa.status,
                   sa.assigned_at,
                   sa.due_date
            FROM speech_assignments sa
            JOIN users u ON u.id = sa.patient_id
            JOIN speech_exercises se ON se.id = sa.exercise_id
            WHERE sa.therapist_id = %s
            ORDER BY sa.assigned_at DESC
            LIMIT 20
        """, (therapist_id,))

        rows = cur.fetchall()
        cur.close()

        out = []
        for r in rows:
            out.append({
                "id": r[0],
                "patient_id": r[1],
                "patient_username": r[2],
                "exercise_title": r[3],
                "status": r[4],
                "assigned_at": r[5].strftime("%Y-%m-%d %H:%M:%S") if r[5] else "",
                "due_date": r[6].strftime("%Y-%m-%d") if r[6] else ""
            })

        return jsonify({"success": True, "assignments": out}), 200

    except Exception as e:
        print("api_speech_recent_assignments error:", e)
        return jsonify({"success": False, "error": "Failed to fetch assignments"}), 500


@app.route("/api/patient/speech_assignments", methods=["GET"])
def api_patient_speech_assignments():
    # patient must be logged in
    if "username" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    identifier = session.get("username")  # username OR email
    try:
        cur = mysql.connection.cursor()

        # get user_id + role
        cur.execute("""
            SELECT id, role
            FROM users
            WHERE username=%s OR email=%s
            LIMIT 1
        """, (identifier, identifier))
        urow = cur.fetchone()
        if not urow:
            cur.close()
            return jsonify({"success": False, "error": "User not found"}), 404

        user_id = urow[0]
        role = (urow[1] or "patient").lower()

        # only patients can access
        if role != "patient":
            cur.close()
            return jsonify({"success": False, "error": "Forbidden"}), 403

        # fetch assignments for this patient
        cur.execute("""
            SELECT
                sa.id,
                se.title,
                se.instructions,
                sa.status,
                sa.assigned_at,
                sa.due_date
            FROM speech_assignments sa
            JOIN speech_exercises se ON se.id = sa.exercise_id
            WHERE sa.patient_id = %s
            ORDER BY sa.assigned_at DESC
            LIMIT 50
        """, (user_id,))
        rows = cur.fetchall()
        cur.close()

        out = []
        for r in rows:
            out.append({
                "id": r[0],
                "title": r[1],
                "instructions": r[2] or "",
                "status": r[3] or "assigned",
                "assigned_at": r[4].strftime("%Y-%m-%d %H:%M:%S") if r[4] else "",
                "due_date": r[5].strftime("%Y-%m-%d") if r[5] else ""
            })

        return jsonify({"success": True, "assignments": out}), 200

    except Exception as e:
        print("api_patient_speech_assignments error:", e)
        return jsonify({"success": False, "error": "Failed to fetch assignments"}), 500

@app.route('/delete_account', methods=['POST'])
def delete_account():
    if 'username' not in session:
        flash('Please log in first.')
        return redirect(url_for('login'))
    
    username = session['username']
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM users WHERE username = %s", (username,))
    mysql.connection.commit()
    cur.close()
    session.pop('username', None)
    flash('Account deleted successfully.')
    return redirect(url_for('home'))


@app.route('/')
def home():
    role = None

    if 'username' in session:
        identifier = session.get('username')  # username OR email
        cur = mysql.connection.cursor()
        cur.execute("""
            SELECT role
            FROM users
            WHERE username=%s OR email=%s
            LIMIT 1
        """, (identifier, identifier))
        row = cur.fetchone()
        cur.close()

        role = (row[0] if row and row[0] else "patient").lower()
        return render_template('home.html', username=identifier, role=role)

    return render_template('home.html', role=None)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        identifier = request.form['identifier']
        password = request.form['password']
        
        # Add restriction check: reject if identifier has digits or '1234'
        if any(char.isdigit() for char in identifier):
            flash('Username must not contain numbers')
            return redirect(url_for('login'))

        cur = mysql.connection.cursor()
        cur.execute("SELECT username, password FROM login WHERE username = %s OR email = %s", (identifier, identifier))
        user = cur.fetchone()
        cur.close()
        if user and bcrypt.check_password_hash(user[1], password):
            session['username'] = user[0]
            # Clear any leftover files from previous session
            user_upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], user[0])
            if os.path.exists(user_upload_dir):
                for file_name in os.listdir(user_upload_dir):
                    file_path = os.path.join(user_upload_dir, file_name)
                    try:
                        if os.path.isfile(file_path):
                            os.remove(file_path)
                    except Exception as e:
                        print(f"Failed to delete {file_path}. Reason: {e}")
            return redirect(url_for('home'))
        else:
            flash('Invalid credentials. Please try again.')
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        if len(password) < 8 or not any(c.isupper() for c in password) or not any(c.isdigit() for c in password):
            flash('Password must be at least 8 characters, with an uppercase letter and a number.')
            return redirect(url_for('register'))
        hashed_pwd = bcrypt.generate_password_hash(password).decode('utf-8')
        cur = mysql.connection.cursor()
        try:
            cur.execute("INSERT INTO login (username, email, password) VALUES (%s, %s, %s)", (username, email, hashed_pwd))
            mysql.connection.commit()
            session['username'] = username
            flash('Registration successful!')
            return redirect(url_for('home'))
        except Exception as e:
            mysql.connection.rollback()
            flash('Username or email already exists.')
            return redirect(url_for('register'))
        finally:
            cur.close()
    return render_template('register.html')

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form['email']
        cur = mysql.connection.cursor()
        cur.execute("SELECT username FROM login WHERE email = %s", (email,))
        user = cur.fetchone()
        cur.close()
        if user:
            reset_token = 'example_reset_token'
            print(f"Reset link for {user[0]}: http://yourdomain.com/reset_password?token={reset_token}")
            flash('Password reset link sent to your email.')
        else:
            flash('Email not found.')
        return redirect(url_for('login'))
    return render_template('forgot_password.html')

@app.route('/logout')
def logout():
    session.pop('username', None)             # Remove your stored email/username
    session.pop('google_oauth_token', None)     # Remove the OAuth token from session
    flash('You have been logged out.')
    return redirect(url_for('home'))


import json

def get_users():
    """Retrieve all user records from the file system."""
    users = []
    for filename in os.listdir(app.config['UPLOAD_FOLDER']):
        if filename.endswith('.json'):
            with open(os.path.join(app.config['UPLOAD_FOLDER'], filename), 'r') as f:
                user = json.load(f)
                users.append(user)
    return users

def save_user(user_data):
    """Save a user record as a JSON file. A simple ID is generated if not provided."""
    user_id = user_data.get('id')
    if not user_id:
        user_id = str(len(os.listdir(app.config['UPLOAD_FOLDER'])) + 1)
        user_data['id'] = user_id
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{user_id}.json")
    with open(filepath, 'w') as f:
        json.dump(user_data, f)

def delete_user(user_id):
    """Delete a user's JSON file based on user ID."""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{user_id}.json")
    if os.path.exists(filepath):
        os.remove(filepath)

def archive_users():
    """Archive all user files into a zip archive and return its path."""
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            zipf.write(filepath, arcname=filename)
    zip_buffer.seek(0)
    archive_name = 'user_archive.zip'
    archive_path = os.path.join(app.config['ARCHIVE_FOLDER'], archive_name)
    with open(archive_path, 'wb') as f:
        f.write(zip_buffer.read())
    return archive_path

# Updated manage_user route
@app.route('/manage_user')
def manage_user():
    audio_files = []
    upload_dir = app.config['UPLOAD_DIR']

    for filename in os.listdir(upload_dir):
        file_path = os.path.join(upload_dir, filename)

        if os.path.isfile(file_path) and filename.lower().endswith(('.mp3', '.wav', '.ogg', '.m4a')):
            ts = os.path.getmtime(file_path)
            audio_files.append({
                'filename': filename,
                'filepath': file_path,
                'upload_time': datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
            })

    return render_template('manage_user.html', audio_files=audio_files)

# 
# Updated archive route
@app.route('/archive')
def archive():
    """Archive all audio files into a zip file"""
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for filename in os.listdir(app.config['UPLOAD_DIR']):
            file_path = os.path.join(app.config['UPLOAD_DIR'], filename)
            if os.path.isfile(file_path):
                zipf.write(file_path, arcname=filename)
    
    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name='audio_archive.zip',
        mimetype='application/zip'
    )
# Updated export route
@app.route('/export')
def export():
    """Export list of all audio files as JSON"""
    audio_data = []
    for username in os.listdir(app.config['UPLOAD_DIR']):
        user_dir = os.path.join(app.config['UPLOAD_DIR'], username)
        if os.path.isdir(user_dir):
            for filename in os.listdir(user_dir):
                file_path = os.path.join(user_dir, filename)
                audio_data.append({
                    'username': username,
                    'filename': filename,
                    'size': os.path.getsize(file_path),
                    'upload_time': os.path.getmtime(file_path)
                })
    
    return send_file(
        BytesIO(json.dumps(audio_data, indent=2).encode()),
        as_attachment=True,
        download_name='audio_files_export.json',
        mimetype='application/json'
    )
@app.route('/import', methods=['POST'])
def import_users():
    file = request.files.get('import_file')
    if file:
        try:
            data = json.load(file)
            # Data should be a list of user objects
            if isinstance(data, list):
                for user in data:
                    save_user(user)
                flash('Users imported successfully!')
            else:
                flash('Invalid file format. Expected a list of users.')
        except Exception as e:
            flash('Error importing users: ' + str(e))
    else:
        flash('No file selected!')
    return redirect(url_for('manage_user'))




WHISPER_MODEL = whisper.load_model("base")

FILLERS_SINGLE = {
    "um", "uh", "erm", "hmm", "ah", "eh", "mm", "mhm", "uhh", "umm"
}

FILLER_PHRASES = [
    "you know",
    "i mean",
    "kind of",
    "sort of",
]

# ✅ add extra common spoken fillers (easy + effective)
FILLERS_SINGLE.update({
    "like", "actually", "basically", "literally", "okay", "right", "well",
    "just", "so"
})

# ✅ extra common phrases
FILLER_PHRASES.extend([
    "to be honest",
    "to be fair",
    "i guess",
    "you see",
])

def fluency_rewrite(text: str) -> str:
    """
    Light, rule-based rewrite to improve fluency while keeping meaning.
    This will change the transcript more often than basic filler removal.
    """
    if not text or not text.strip():
        return ""

    t = text.strip()

    # 1) Remove filler phrases
    for p in FILLER_PHRASES:
        t = re.sub(rf"\b{re.escape(p)}\b", "", t, flags=re.IGNORECASE)

    # 2) Remove letter stutters: b-b-b-but / t-t-the
    t = re.sub(r"\b([a-zA-Z])(?:-\1){1,10}-([a-zA-Z]+)\b", r"\2", t)

    # 3) Collapse repeated starters even with punctuation between them
    # "so... so", "okay, okay", "well - well"
    t = re.sub(r"\b(so|okay|well|right)\b(?:\s*[\.,!\?\-–—]+\s*|\s+)\1\b", r"\1", t, flags=re.IGNORECASE)

    # 4) Token pass: remove single-word fillers + immediate duplicates
    tokens = t.split()
    cleaned = []
    prev = None
    for tok in tokens:
        n = _norm(tok)
        if n in FILLERS_SINGLE:
            continue
        if prev is not None and n and n == prev:
            continue
        cleaned.append(tok)
        prev = n

    t = " ".join(cleaned)

    # 5) Remove short false-start fragments: "the - the" / "i - i"
    t = re.sub(r"\b(\w+)\s*[-–—]\s*\1\b", r"\1", t, flags=re.IGNORECASE)

    # 6) Cleanup spacing
    t = re.sub(r"\s+([,.!?;:])", r"\1", t)
    t = re.sub(r"\s{2,}", " ", t).strip()

    return t if t else text.strip()


def _norm(tok: str) -> str:
    return re.sub(r"[^a-z0-9']+", "", tok.lower())

def remove_disfluencies(text: str) -> str:
    """
    Removes:
    - filler words: um/uh/erm...
    - filler phrases: you know / i mean ...
    - immediate repeats: I I I -> I
    - repeated bigrams: I want I want -> I want
    - letter stutters: b-b-b-but -> but
    """
    if not text or not text.strip():
        return ""

    t = text.strip()

    # remove filler phrases
    for p in FILLER_PHRASES:
        t = re.sub(rf"\b{re.escape(p)}\b", "", t, flags=re.IGNORECASE)

    # remove letter-stutter patterns: b-b-but / t-t-the
    t = re.sub(r"\b([a-zA-Z])(?:-\1){1,10}-([a-zA-Z]+)\b", r"\2", t)

    tokens = t.split()

    # remove filler single words + immediate duplicates
    cleaned = []
    prev = None
    for tok in tokens:
        n = _norm(tok)
        if n in FILLERS_SINGLE:
            continue
        if prev is not None and n and n == prev:
            continue
        cleaned.append(tok)
        prev = n

    # remove repeated bigrams (A B A B)
    norms = [_norm(x) for x in cleaned]
    out = []
    i = 0
    while i < len(cleaned):
        if i + 3 < len(cleaned):
            a1, b1 = norms[i], norms[i+1]
            a2, b2 = norms[i+2], norms[i+3]
            if a1 and b1 and a1 == a2 and b1 == b2:
                out.append(cleaned[i])
                out.append(cleaned[i+1])
                i += 4
                continue
        out.append(cleaned[i])
        i += 1

    final = " ".join(out)
    final = re.sub(r"\s+([,.!?;:])", r"\1", final)
    final = re.sub(r"\s{2,}", " ", final).strip()
    return final if final else text.strip()

import subprocess

def to_wav16k_mono(src_path: str, dst_wav: str):
    """
    Convert any audio (ogg/webm/mp3/wav) to 16kHz mono wav using our ffmpeg exe.
    """
    cmd = [
        FFMPEG_EXE, "-y",
        "-i", src_path,
        "-ar", "16000",
        "-ac", "1",
        dst_wav
    ]
    subprocess.run(cmd, check=True, capture_output=True)

def _save_upload_to_temp(uploaded_file, suffix):
    fd, path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    uploaded_file.save(path)
    return path

@app.route("/convert_to_wav", methods=["POST"])
def convert_to_wav():
    if "file" not in request.files:
        return {"success": False, "error": "No file"}, 400

    f = request.files["file"]
    in_path = _save_upload_to_temp(f, ".webm")
    out_fd, out_path = tempfile.mkstemp(suffix=".wav")
    os.close(out_fd)

    try:
        subprocess.run([
            FFMPEG_EXE, "-y",
            "-i", in_path,
            "-vn",
            "-acodec", "pcm_s16le",
            "-ar", "44100",
            "-ac", "1",
            out_path
        ], check=True)

        return send_file(out_path, as_attachment=True,
                         download_name="recorded.wav",
                         mimetype="audio/wav")
    finally:
        try: os.remove(in_path)
        except: pass

@app.route("/convert_to_mp3", methods=["POST"])
def convert_to_mp3():
    if "file" not in request.files:
        return {"success": False, "error": "No file"}, 400

    f = request.files["file"]
    in_path = _save_upload_to_temp(f, ".webm")
    out_fd, out_path = tempfile.mkstemp(suffix=".mp3")
    os.close(out_fd)

    try:
        subprocess.run([
            FFMPEG_EXE, "-y",
            "-i", in_path,
            "-vn",
            "-codec:a", "libmp3lame",
            "-b:a", "192k",
            out_path
        ], check=True)

        return send_file(out_path, as_attachment=True,
                         download_name="recorded.mp3",
                         mimetype="audio/mpeg")
    finally:
        try: os.remove(in_path)
        except: pass

def whisper_transcribe_en(audio_path: str) -> str:
    """
    Transcribe forcing English output.
    """
    result = WHISPER_MODEL.transcribe(audio_path, language="en", fp16=False)
    return (result.get("text") or "").strip()

async def _edge_tts_save(text: str, out_mp3: str, voice: str):
    communicate = edge_tts.Communicate(text, voice=voice)
    await communicate.save(out_mp3)

def tts_edge_mp3(text: str, out_mp3: str, voice: str = "en-US-JennyNeural"):
    asyncio.run(_edge_tts_save(text, out_mp3, voice))


def TEXT_CORRECTION(text):
    """
    Corrects text using TextBlob.
    """
    # Download NLTK resources quietly (only needed once)
    nltk.download('punkt', quiet=True)
    corrected_text = TextBlob(text).correct()
    return str(corrected_text)

def text_to_audio(text):
    """
    Dummy text-to-speech conversion.
    (In a real scenario, integrate with a TTS engine.)
    This function simply writes dummy data to simulate output.
    """
    output_file = "regen_audio.mp3"
    # For demonstration, write some bytes to simulate an audio file.
    with open(output_file, 'wb') as f:
        f.write(b"FAKE_AUDIO_DATA")
    return output_file

def enhance_audio(audio_path):
    """
    Process (or 'enhance') the audio by converting it to text,
    performing text correction and then converting it back to audio.
    Returns the path of the final (simulated) audio, the corrected text, and the raw text.
    """
    # Step 1: Audio to Text conversion
    raw_text = audio_to_text(audio_path)
    # Step 2: Correct the text
    corrected_text = TEXT_CORRECTION(raw_text)
    # Step 3: Convert the corrected text back to audio
    final_audio_path = text_to_audio(corrected_text)
    return final_audio_path, corrected_text, raw_text

# -------------------------------
# Routes
# -------------------------------

@app.route('/ui')
def ui():
    if 'username' not in session:
        return redirect(url_for('login'))

    identifier = session.get("username")
    user_id = get_or_create_users_id(identifier)

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id, original_filename, stored_filename, uploaded_at, status
        FROM uploaded_audio
        WHERE user_id = %s
        ORDER BY uploaded_at DESC
    """, (user_id,))
    rows = cur.fetchall()
    cur.close()

    # convert tuples -> dicts
    audios = []
    for r in rows:
        audios.append({
            "id": r[0],
            "original_filename": r[1],
            "stored_filename": r[2],
            "uploaded_at": r[3],
            "status": r[4],
        })

    return render_template('ui.html', audios=audios)


@app.route('/patient_record')
def patient_record():
    patient = {"id": 1, "name": "John Doe"}
    meds = []; speech_logs = []; therapists = []; treatments = []
    return render_template(
        'patient_record.html',
        patient=patient,
        meds=meds,
        speech_logs=speech_logs,
        therapists=therapists,
        treatments=treatments
    )

# ✅ create an endpoint literally named "therapist.invite"
@app.route('/therapist/invite/<int:patient_id>', endpoint='therapist.invite')
def _therapist_invite(patient_id):
    # put your invite logic here if you like
    flash('Invite triggered.')
    # redirect somewhere valid you already have:
    return redirect(url_for('therapist.therapist_notes', patient_id=patient_id))
# app.py (or routes.py)
from flask import render_template, request, redirect, url_for, flash


@app.route('/therapist_dashboard', methods=['GET', 'POST'])
def therapist_dashboard():
    # ✅ must be therapist
    if not _is_therapist_logged_in():
        flash("Unauthorized (therapist only).", "danger")
        return redirect(url_for("home"))

    cur = mysql.connection.cursor()

    # ✅ load real patients from users table
    cur.execute("""
        SELECT id, username, email
        FROM users
        WHERE role = 'patient'
        ORDER BY username ASC
    """)
    patients_rows = cur.fetchall()

    patients = []
    for r in patients_rows:
        patients.append({
            "id": r[0],
            "name": r[1],   # UI uses name
            "email": r[2]
        })

    # ✅ find therapist_id
    identifier = session.get("username")
    cur.execute("""
        SELECT id
        FROM users
        WHERE username=%s OR email=%s
        LIMIT 1
    """, (identifier, identifier))
    trow = cur.fetchone()
    therapist_id = trow[0] if trow else None

    # ✅ Load saved notes for this therapist (LATEST FIRST)
    notes = []
    if therapist_id:
        cur.execute("""
            SELECT tn.id, tn.patient_id, u.username AS patient_name,
                   tn.subject, tn.message, tn.created_at
            FROM therapist_notes tn
            JOIN users u ON u.id = tn.patient_id
            WHERE tn.therapist_id = %s
            ORDER BY tn.created_at DESC
            LIMIT 50
        """, (therapist_id,))
        rows = cur.fetchall()

        for r in rows:
            notes.append({
                "id": r[0],
                "patient_id": r[1],
                "patient_name": r[2],
                "subject": r[3],
                "message": r[4],
                "created_at": r[5].strftime("%Y-%m-%d %H:%M") if r[5] else ""
            })

    # keep your coping_plans sample or replace later
    coping_plans = []

    # ✅ Handle POST only if you still want non-AJAX form posts
    # (If you're using AJAX to /therapist/notes/save, you can remove this whole POST block)
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'save_note':
            patient_id = (request.form.get("patient_id") or "").strip()
            subject    = (request.form.get("subject") or "").strip()
            message    = (request.form.get("message") or "").strip()

            if not patient_id or not subject or not message:
                flash("All fields are required.", "danger")
                cur.close()
                return redirect(url_for("therapist_dashboard"))

            # ✅ verify patient
            cur.execute("SELECT role FROM users WHERE id=%s LIMIT 1", (patient_id,))
            prow = cur.fetchone()
            if not prow or (prow[0] or "").lower() != "patient":
                flash("Selected user is not a patient.", "danger")
                cur.close()
                return redirect(url_for("therapist_dashboard"))

            # ✅ insert note
            cur.execute("""
                INSERT INTO therapist_notes (therapist_id, patient_id, subject, message, created_at)
                VALUES (%s, %s, %s, %s, NOW())
            """, (therapist_id, patient_id, subject, message))
            mysql.connection.commit()

            flash("Secure note saved.", "success")
            cur.close()
            return redirect(url_for("therapist_dashboard"))

    cur.close()

    return render_template(
        'therapist_dashboard.html',
        patients=patients,
        coping_plans=coping_plans,
        notes=notes,
        therapist_id=therapist_id
    )


@app.route("/therapist/report/<int:patient_id>/generate", methods=["POST"])
def therapist_generate_report(patient_id):
    # ✅ Only therapist
    if not _is_therapist_logged_in():
        flash("Unauthorized (therapist only).", "danger")
        return redirect(url_for("home"))

    report_type = (request.form.get("report_type") or "summary").strip()
    from_date   = (request.form.get("from_date") or "").strip()
    to_date     = (request.form.get("to_date") or "").strip()

    # ✅ NEW FIELD (from your textarea)
    therapist_notes = (request.form.get("therapist_notes") or "").strip()
    if not therapist_notes:
        flash("Therapist report text is required.", "danger")
        return redirect(url_for("therapist_dashboard") + "#report")

    # ✅ fetch patient basic info (optional but recommended)
    cur = mysql.connection.cursor()
    cur.execute("SELECT username, email FROM users WHERE id=%s AND role='patient' LIMIT 1", (patient_id,))
    prow = cur.fetchone()
    cur.close()

    if not prow:
        flash("Patient not found.", "danger")
        return redirect(url_for("therapist_dashboard") + "#report")

    patient_username, patient_email = prow[0], prow[1]

    # ✅ Build PDF in memory (ReportLab)
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Patient Report", styles["Title"]))
    story.append(Spacer(1, 8))

    story.append(Paragraph(f"<b>Patient:</b> {patient_username}", styles["Normal"]))
    story.append(Paragraph(f"<b>Email:</b> {patient_email}", styles["Normal"]))
    story.append(Paragraph(f"<b>Report Type:</b> {report_type}", styles["Normal"]))

    if from_date or to_date:
        story.append(Paragraph(f"<b>Date Range:</b> {from_date or '—'} to {to_date or '—'}", styles["Normal"]))

    story.append(Spacer(1, 12))

    # ✅ NEW: Therapist written report section
    story.append(Paragraph("Therapist Report", styles["Heading3"]))
    story.append(Spacer(1, 6))
    story.append(Paragraph(therapist_notes.replace("\n", "<br/>"), styles["BodyText"]))
    story.append(Spacer(1, 12))

    doc.build(story)
    buffer.seek(0)

    filename = f"report_patient_{patient_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")




@app.route('/audio_upload', methods=['POST'])
def audio_upload():
    if 'username' not in session:
        flash("Please login first.")
        return redirect(url_for('login'))

    if 'file' not in request.files:
        flash("No file provided.")
        return redirect(url_for('ui'))

    file = request.files['file']
    if file.filename == '':
        flash("Empty filename.")
        return redirect(url_for('ui'))

    username = session['username']
    user_id = get_user_id(username)

    if not user_id:
        flash("User not found.")
        return redirect(url_for('ui'))

    # ---------- FILE SAVE ----------
    original_filename = file.filename
    safe_name = secure_filename(file.filename)

    user_dir = os.path.join(app.config['UPLOAD_DIR'], username)
    os.makedirs(user_dir, exist_ok=True)

    stored_filename = f"{uuid.uuid4().hex}_{safe_name}"
    file_path = os.path.join(user_dir, stored_filename)

    file.save(file_path)

    # ---------- DB INSERT ----------
    cur = mysql.connection.cursor()
    cur.execute("""
        INSERT INTO uploaded_audio
        (user_id, original_filename, stored_filename, file_path)
        VALUES (%s, %s, %s, %s)
    """, (user_id, original_filename, stored_filename, file_path))

    mysql.connection.commit()
    cur.close()

    flash("Audio uploaded successfully.")
    return redirect(url_for('ui'))

@app.route("/audio/<int:audio_id>")
def stream_audio(audio_id):
    if "username" not in session:
        abort(401)

    identifier = session.get("username")
    user_id = get_or_create_users_id(identifier)  # or your final helper

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT file_path, stored_filename
        FROM uploaded_audio
        WHERE id = %s AND user_id = %s
    """, (audio_id, user_id))
    row = cur.fetchone()
    cur.close()

    if not row:
        abort(404)

    file_path, stored_filename = row

    # send_file can serve absolute paths too
    return send_file(file_path, as_attachment=False, download_name=stored_filename)


@app.route("/upload_recorded", methods=["POST"])
def upload_recorded():
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "error": "No file provided"}), 400

        f = request.files["file"]
        if not f or f.filename == "":
            return jsonify({"success": False, "error": "Empty filename"}), 400

        filename = secure_filename(f.filename)
        save_path = os.path.join(app.config["UPLOAD_DIR"], filename)
        f.save(save_path)

        return jsonify({"success": True, "filename": filename}), 200

    except Exception as e:
        print("upload_recorded error:", e)
        return jsonify({"success": False, "error": "Failed to upload recorded audio"}), 500


@app.route('/convert_text', methods=['POST'])
def convert_text():
    """
    Converts all uploaded audio files to text and displays the result.
    """
    files = os.listdir(app.config['UPLOAD_DIR'])
    audio_texts = {}
    for filename in files:
        file_path = os.path.join(app.config['UPLOAD_DIR'], filename)
        try:
            text = audio_to_text(file_path)
            audio_texts[filename] = text
            flash(f'Converted "{filename}" to text.')
        except Exception as e:
            flash(f'Error converting "{filename}": {str(e)}')
    return render_template('ui.html', files=files, audio_texts=audio_texts)

@app.route('/process_results', methods=['POST'])
def process_results():
    """
    Processes (or "reconstructs") each uploaded audio file.
    The processed audio is copied to a static folder and result details are shown.
    """
    files = os.listdir(app.config['UPLOAD_DIR'])
    result_data = []
    for filename in files:
        file_path = os.path.join(app.config['UPLOAD_DIR'], filename)
        try:
            final_audio_path, corrected_text, raw_text = enhance_audio(file_path)
            # Copy the processed audio to a static folder for playback
            processed_dir = os.path.join(app.static_folder, 'processed')
            os.makedirs(processed_dir, exist_ok=True)
            dest_filename = f"processed_{filename}"
            destination = os.path.join(processed_dir, dest_filename)
            shutil.copy(final_audio_path, destination)
            result_data.append({
                'filename': filename,
                'final_audio': url_for('static', filename='processed/' + dest_filename),
                'corrected_text': corrected_text,
                'raw_text': raw_text
            })
            flash(f'Processed "{filename}" successfully.')
        except Exception as e:
            flash(f'Error processing "{filename}": {str(e)}')
    return render_template('ui.html', files=files, result_data=result_data)

from gtts import gTTS

def tts_gtts_mp3(text: str, out_mp3: str):
    tts = gTTS(text=text, lang="en")
    tts.save(out_mp3)

@app.route("/stutter_remove", methods=["POST"])
def stutter_remove():
    import os, uuid, tempfile, traceback
    from werkzeug.utils import secure_filename

    in_path = None
    out_path = None

    try:
        if "file" not in request.files:
            return jsonify({"error": "No file part in request"}), 400

        f = request.files["file"]
        if not f or f.filename == "":
            return jsonify({"error": "No file selected"}), 400

        with tempfile.TemporaryDirectory() as td:
            in_path = os.path.join(td, secure_filename(f.filename))
            f.save(in_path)
            print("1) saved file ->", in_path)

            # --- ASR ---
            wav_path = os.path.join(td, f"in_{uuid.uuid4().hex}.wav")
            print("2) converting to wav ->", wav_path)
            to_wav16k_mono(in_path, wav_path)
            print("3) convert ok, exists =", os.path.exists(wav_path))

            raw_text = whisper_transcribe_en(wav_path)
            print("4) whisper ok ->", raw_text)

            # --- CLEAN TEXT ---
            clean_text = fluency_rewrite(raw_text)
            print("3) cleaned ok ->", clean_text)

            if not clean_text.strip():
                return jsonify({"error": "Empty transcription", "raw_text": raw_text}), 422

            # --- TTS (choose ONE) ---
            # Option A: gTTS (recommended quick win)
            out_path = os.path.join(td, f"clean_{uuid.uuid4().hex}.mp3")
            print("4) starting gTTS ->", out_path)
            tts_gtts_mp3(clean_text, out_path)
            print("5) gTTS done")

            if not os.path.exists(out_path):
                return jsonify({"error": "TTS output not created", "out_path": out_path}), 500

        
            from io import BytesIO

            with open(out_path, "rb") as f:
                audio_bytes = f.read()

            buf = BytesIO(audio_bytes)
            buf.seek(0)

            resp = send_file(
                buf,
                mimetype="audio/mpeg",
                as_attachment=False,
                download_name="cleaned.mp3"
            )
            resp.headers["X-Raw-Text"] = raw_text[:5000]
            resp.headers["X-Clean-Text"] = clean_text[:5000]
            return resp


    except Exception as e:
        traceback.print_exc()
        return jsonify({
            "error": f"{type(e).__name__}: {str(e)}",
            "in_path": in_path,
            "out_path": out_path
        }), 500


@app.route("/stutter_remove_db/<int:audio_id>", methods=["POST"])
def stutter_remove_db(audio_id):
    import os, uuid, tempfile, traceback

    try:
        # 1) Auth (your session uses 'username' as identifier)
        if "username" not in session:
            return jsonify({"success": False, "error": "Unauthorized"}), 401

        identifier = session.get("username")  # could be email if google login
        user_id = get_or_create_users_id(identifier)
        if not user_id:
            return jsonify({"success": False, "error": "User not found"}), 404

        # 2) Fetch original uploaded audio record
        row = db_get_uploaded_audio_by_id(mysql, audio_id, user_id)
        if not row:
            return jsonify({"success": False, "error": "Audio not found"}), 404

        # row = (id, user_id, original_filename, stored_filename, file_path)
        uploaded_audio_id = row[0]
        original_filename = row[2]
        file_path = row[4]

        if not file_path or not os.path.exists(file_path):
            return jsonify({"success": False, "error": "Audio file missing on server"}), 404

        # 3) Run your EXACT existing pipeline in a temp dir
        with tempfile.TemporaryDirectory() as td:
            # Convert to wav 16k mono
            wav_path = os.path.join(td, f"in_{uuid.uuid4().hex}.wav")
            to_wav16k_mono(file_path, wav_path)

            raw_text = whisper_transcribe_en(wav_path)
            clean_text = fluency_rewrite(raw_text)

            if not clean_text.strip():
                return jsonify({"success": False, "error": "Empty transcription", "raw_text": raw_text}), 422

            # TTS output
            cleaned_filename = f"clean_{uuid.uuid4().hex}.mp3"
            out_path = os.path.join(td, cleaned_filename)
            tts_gtts_mp3(clean_text, out_path)

            if not os.path.exists(out_path):
                return jsonify({"success": False, "error": "TTS output not created"}), 500

            # 4) Read cleaned audio bytes (BLOB)
            with open(out_path, "rb") as f:
                cleaned_audio_bytes = f.read()

        # 5) Store in DB
        result_id = db_upsert_stutter_removed(
            mysql=mysql,
            user_id=user_id,
            uploaded_audio_id=uploaded_audio_id,
            original_filename=original_filename,
            cleaned_filename=cleaned_filename,
            cleaned_audio_bytes=cleaned_audio_bytes,
            raw_text=raw_text,
            clean_text=clean_text
        )

        return jsonify({
            "success": True,
            "result_id": result_id,
            "uploaded_audio_id": uploaded_audio_id,
            "original_filename": original_filename,
            "cleaned_filename": cleaned_filename,
            "raw_text": raw_text,
            "clean_text": clean_text
        }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "error": f"{type(e).__name__}: {str(e)}"}), 500


from flask import Response

@app.route("/stutter_result_audio/<int:result_id>", methods=["GET"])
def stutter_result_audio(result_id):
    if "username" not in session:
        abort(401)

    identifier = session.get("username")
    user_id = get_or_create_users_id(identifier)
    if not user_id:
        abort(404)

    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT cleaned_audio, cleaned_filename
        FROM stutter_removed_audio
        WHERE id=%s AND user_id=%s
        LIMIT 1
    """, (result_id, user_id))
    row = cur.fetchone()
    cur.close()

    if not row or row[0] is None:
        abort(404)

    audio_bytes = row[0]
    filename = row[1] or "cleaned.mp3"

    # ✅ check if download is requested
    download = request.args.get("download") == "1"
    disposition = "attachment" if download else "inline"

    return Response(
        audio_bytes,
        mimetype="audio/mpeg",
        headers={
            "Content-Disposition": f'{disposition}; filename="{filename}"'
        }
    )

@app.route("/api/results/audios", methods=["GET"])
def api_results_audios():
    # 1) must be logged in
    if "username" not in session:
        return jsonify({"success": False, "error": "Unauthorized"}), 401

    identifier = session.get("username")
    user_id = get_or_create_users_id(identifier)
    if not user_id:
        return jsonify({"success": False, "error": "User not found"}), 404

    cur = mysql.connection.cursor()

    # 2) uploaded audios
    cur.execute("""
        SELECT id, original_filename, stored_filename, uploaded_at
        FROM uploaded_audio
        WHERE user_id = %s
        ORDER BY uploaded_at DESC
    """, (user_id,))
    up_rows = cur.fetchall()

    uploaded = []
    for r in up_rows:
        uploaded.append({
            "id": r[0],
            "original_filename": r[1],
            "stored_filename": r[2],
            "uploaded_at": r[3].strftime("%Y-%m-%d %H:%M:%S") if r[3] else None,
            # ✅ use your existing stream route
            "listen_url": url_for("stream_audio", audio_id=r[0])
        })

    # 3) stutter removed results
    cur.execute("""
        SELECT id, uploaded_audio_id, original_filename, cleaned_filename, created_at, raw_text, clean_text
        FROM stutter_removed_audio
        WHERE user_id = %s
        ORDER BY created_at DESC
    """, (user_id,))
    sr_rows = cur.fetchall()
    cur.close()

    cleaned = []
    for r in sr_rows:
        cleaned.append({
            "id": r[0],
            "uploaded_audio_id": r[1],
            "original_filename": r[2],
            "cleaned_filename": r[3],
            "created_at": r[4].strftime("%Y-%m-%d %H:%M:%S") if r[4] else None,
            "raw_text": r[5] or "",
            "clean_text": r[6] or "",
            # ✅ use your existing blob-stream route
            "listen_url": url_for("stutter_result_audio", result_id=r[0])
        })

    return jsonify({
        "success": True,
        "uploaded": uploaded,
        "cleaned": cleaned
    }), 200

# For processed audio files, we'll copy them into a subfolder in the static directory.
PROCESSED_DIR = os.path.join(app.static_folder, 'processed')
os.makedirs(PROCESSED_DIR, exist_ok=True)


def audio_to_text(audio_path):
    """Convert audio to text using Whisper."""
    model_whisper = whisper.load_model("base")
    result = model_whisper.transcribe(audio_path)
    return result["text"]

def TEXT_CORRECTION(text):
    """Corrects the input text using TextBlob."""
    nltk.download('punkt', quiet=True)
    corrected_text = TextBlob(text).correct()
    return str(corrected_text)

def text_to_audio(text, username):
    """
    Dummy text-to-speech conversion.
    In a real system, integrate with a proper TTS engine.
    Here we simply write dummy bytes to simulate an output audio file.
    """
    output_file = f'regenerated_{username}.mp3'
    with open(output_file, 'wb') as f:
        f.write(b"FAKE_AUDIO_DATA")
    return output_file

def enhance_audio(audio_path, username):
    """
    Simulate audio enhancement by:
      1. Transcribing the audio.
      2. Correcting the transcription.
      3. Converting the corrected text back to audio.
    Returns the final (dummy) audio file path, the corrected text, and the raw transcription.
    """
    # (For simplicity, we don’t perform real audio processing here.)
    raw_text = audio_to_text(audio_path)
    corrected_text = TEXT_CORRECTION(raw_text)
    final_audio_path = text_to_audio(corrected_text, username)
    return final_audio_path, corrected_text, raw_text

def predict_stutter(audio_file):
    """
    Dummy stutter prediction: returns 1 if the length of the transcription is even,
    otherwise 0.
    """
    text = audio_to_text(audio_file)
    return 1 if len(text) % 2 == 0 else 0


@app.route("/results")
def results():
    if "username" not in session:
        return redirect(url_for("login"))
    return render_template("results.html")


#########  MOdule 6
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file

# feedback_bp = Blueprint('feedback', __name__)


# Global list to store feedback entries (for production, use a database)
feedback_entries = []

# Ensure the folder exists (you can call it "text" or any name you like)
FEEDBACK_FOLDER = 'text'
os.makedirs(FEEDBACK_FOLDER, exist_ok=True)

@app.route('/feedback', methods=['GET', 'POST'])
def feedback():
    if request.method == 'POST':
        feedback_text = request.form.get('feedbackText')
        rating = request.form.get('rating')

        if not feedback_text or not rating:
            flash("Please provide both feedback and rating.")
            return redirect(url_for('feedback'))

        # ✅ Get username from session
        username = session.get("username", None)
        if not username:
            flash("You must be logged in to submit feedback.")
            return redirect(url_for('feedback'))

        today = datetime.now().strftime("%Y-%m-%d")

        # ✅ CHECK: has this user already submitted feedback today?
        for entry in feedback_entries:
            if (
                entry.get("username") == username and
                entry.get("timestamp", "").startswith(today)
            ):
                flash("You can only submit feedback once per day.")
                return redirect(url_for('feedback'))

        # ✅ If not submitted today → allow
        entry = {
            "username": username,
            "feedbackText": feedback_text,
            "rating": int(rating),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        feedback_entries.append(entry)

        filename = f"feedback_{datetime.now().strftime('%Y%m%d%H%M%S')}.txt"
        filepath = os.path.join(FEEDBACK_FOLDER, filename)
        with open(filepath, 'w') as f:
            json.dump(entry, f, indent=4)

        flash("Feedback submitted successfully!")
        return redirect(url_for('feedback'))

    return render_template('feedback.html', feedback_entries=feedback_entries)


@app.route('/export_feedback')
def export_feedback():
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Feedback Report", styles["Title"]))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]))
    story.append(Spacer(1, 12))

    if not feedback_entries:
        story.append(Paragraph("No feedback submitted yet.", styles["Normal"]))
        doc.build(story)
        buffer.seek(0)
        return send_file(buffer, as_attachment=True,
                         download_name="feedback_report.pdf",
                         mimetype="application/pdf")

    # ✅ Added "Username" column
    table_data = [["#", "Username", "Rating", "Timestamp", "Feedback"]]

    for i, entry in enumerate(feedback_entries, start=1):
        table_data.append([
            str(i),
            entry.get("username", "Unknown"),
            str(entry.get("rating", "")),
            entry.get("timestamp", ""),
            entry.get("feedbackText", "")
        ])

    # Adjust widths because we added a new column
    table = Table(table_data, colWidths=[25, 90, 50, 120, 265])

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    story.append(table)
    doc.build(story)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True,
                     download_name="feedback_report.pdf",
                     mimetype="application/pdf")



@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """
    Serves an uploaded file.
    """
    return send_from_directory(app.config['UPLOAD_DIR'], filename)

@app.route('/delete_audio/<filename>', methods=['POST'])
def delete_audio_ui(filename):
    try:
        # security: stop path traversal like ../../
        safe_name = os.path.basename(filename)

        file_path = os.path.join(app.config['UPLOAD_DIR'], safe_name)

        if not os.path.exists(file_path):
            return jsonify({"success": False, "error": "File not found"}), 404

        os.remove(file_path)
        return jsonify({"success": True}), 200

    except Exception as e:
        print("Delete error:", e)
        return jsonify({"success": False, "error": "Delete failed"}), 500



@app.route('/delete_uploaded_audio/<int:audio_id>', methods=['POST'])
def delete_uploaded_audio(audio_id):
    try:
        # ✅ Use whatever you store in session for login
        identifier = session.get('email') or session.get('username')
        if not identifier:
            return jsonify({"success": False, "error": "Not logged in"}), 401

        cur = mysql.connection.cursor()

        # ✅ Find logged-in user id
        cur.execute("SELECT id FROM users WHERE email=%s OR username=%s", (identifier, identifier))
        user_row = cur.fetchone()
        if not user_row:
            cur.close()
            return jsonify({"success": False, "error": "User not found"}), 404

        user_id = user_row[0]

        # ✅ Only delete if this audio belongs to this user
        cur.execute("SELECT id FROM uploaded_audio WHERE id=%s AND user_id=%s", (audio_id, user_id))
        owned = cur.fetchone()
        if not owned:
            cur.close()
            return jsonify({"success": False, "error": "Audio not found or not allowed"}), 404

        # ✅ Delete DB row
        cur.execute("DELETE FROM uploaded_audio WHERE id=%s AND user_id=%s", (audio_id, user_id))
        mysql.connection.commit()
        cur.close()

        return jsonify({"success": True}), 200

    except Exception as e:
        print("delete_uploaded_audio error:", e)
        return jsonify({"success": False, "error": "Delete failed"}), 500

@app.route("/parent_monitor", methods=["GET"], strict_slashes=False)
def parent_monitor():
    parent_id = session.get("parent_id", 1)
    children = [{"id":101,"name":"Ayaan"},{"id":202,"name":"Hiba"},{"id":303,"name":"Musa"}]

    child_id = request.args.get("child_id", type=int) or children[0]["id"]
    selected_child = next((c for c in children if c["id"] == child_id), children[0])

    kpis = {"completed":18,"pending":6,"overdue":1,"focus_minutes":245,"points":320}

    today = datetime.now().date()  # <-- class datetime (OK)
    series_values = [35,42,28,55,60,15,10]
    last7 = [{"label": (today - timedelta(days=i)).strftime("%a"), "minutes": v}
             for i, v in enumerate(series_values[::-1])]
    chart_labels = [d["label"] for d in last7]
    chart_series = [d["minutes"] for d in last7]

    chart_labels_json = json.dumps(chart_labels)
    chart_series_json = json.dumps(chart_series)

    upcoming = [
        {"title":"Math worksheet","due":(today+timedelta(days=1)).strftime("%b %d, %Y"),"priority":"Normal"},
        {"title":"Read 10 pages","due":(today+timedelta(days=2)).strftime("%b %d, %Y"),"priority":"Normal"},
        {"title":"Science quiz prep","due":(today+timedelta(days=3)).strftime("%b %d, %Y"),"priority":"High"},
    ]
    completed = [
        {"title":"Finish Chapter 1","done_on":(today).strftime("%b %d, %Y")},
        {"title":"30-min focus session","done_on":(today - timedelta(days=1)).strftime("%b %d, %Y")},
        {"title":"Vocabulary practice","done_on":(today - timedelta(days=2)).strftime("%b %d, %Y")},
    ]

    return render_template(
        "parent_monitor.html",
        parent_id=parent_id,
        children=children,
        selected_child=selected_child,
        kpis=kpis,
        chart_labels_json=chart_labels_json,
        chart_series_json=chart_series_json,
        upcoming=upcoming,
        completed=completed,
        view_report_url=url_for("view_child_report", child_id=child_id),
        export_pdf_url=url_for("export_child_report_pdf", child_id=child_id),
    )

# --- tiny placeholders so your buttons don't 404 ---
@app.route("/report/<int:child_id>")
def view_child_report(child_id):
    flash(f"Viewing report for child {child_id} (placeholder).")
    return redirect(url_for("parent_monitor", child_id=child_id))

@app.route("/report/<int:child_id>/pdf")
def export_child_report_pdf(child_id):
    flash(f"Exporting PDF for child {child_id} (placeholder).")
    return redirect(url_for("parent_monitor", child_id=child_id))



@app.route('/enhance_audio', methods=['POST'])
def enhance_audio_route():
    if 'file' not in request.files:
        return jsonify({"error": "No file part in request"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    audio_bytes = file.read()

    try:
        # ✅ use the uniquely named function
        enhanced_audio, sr = autotune_enhance_audio(io.BytesIO(audio_bytes))

        buf = io.BytesIO()
        sf.write(buf, enhanced_audio, sr, format='WAV')
        buf.seek(0)

        return send_file(
            buf,
            mimetype='audio/wav',
            as_attachment=True,
            download_name='enhanced_audio.wav'
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"{type(e).__name__}: {str(e)}"}), 500

    


@app.route("/user_management")
def user_management_page():
    q = (request.args.get("q") or "").strip()

    cur = mysql.connection.cursor()

    if q:
        like = f"%{q}%"
        cur.execute(
            """
            SELECT id, email, password, full_name, username, dob, bio, phone, address
            FROM `login`.`users`
            WHERE username LIKE %s OR email LIKE %s
            ORDER BY id ASC
            """,
            (like, like),
        )
    else:
        cur.execute(
            """
            SELECT id, email, password, full_name, username, dob, bio, phone, address
            FROM `login`.`users`
            ORDER BY id ASC
            """
        )

    users = cur.fetchall()
    cur.close()

    return render_template("user_management.html", users=users, q=q)


@app.route("/delete_user/<int:user_id>", methods=["POST"])
def delete_user(user_id):
    cur = mysql.connection.cursor()
    cur.execute("DELETE FROM `login`.`users` WHERE id = %s", (user_id,))
    mysql.connection.commit()
    cur.close()

    flash("User deleted successfully.", "success")
    return redirect(url_for("user_management_page"))


@app.route("/edit_user", methods=["POST"])
def edit_user():
    user_id   = request.form.get("id")
    full_name = request.form.get("full_name")
    username  = request.form.get("username")
    email     = request.form.get("email")
    dob       = request.form.get("dob")        # can be "" -> set to NULL
    bio       = request.form.get("bio")
    phone     = request.form.get("phone")
    address   = request.form.get("address")

    if not email:
        flash("Email is required.", "danger")
        return redirect(url_for("user_management_page"))

    # Convert empty DOB to NULL (prevents date format errors)
    dob = dob.strip() if dob else ""
    dob_value = dob if dob else None

    cur = mysql.connection.cursor()
    cur.execute("""
        UPDATE `login`.`users`
        SET
            full_name = %s,
            username  = %s,
            email     = %s,
            dob       = %s,
            bio       = %s,
            phone     = %s,
            address   = %s
        WHERE id = %s
    """, (full_name, username, email, dob_value, bio, phone, address, user_id))

    mysql.connection.commit()
    cur.close()

    flash("User updated successfully.", "success")
    return redirect(url_for("user_management_page"))



@app.route("/export_users_excel")
def export_users_excel():
    # 1) Fetch all columns from users table (same as UI)
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT id, email, password, full_name, username, dob, bio, phone, address
        FROM `login`.`users`
        ORDER BY id ASC
    """)
    rows = cur.fetchall()
    cur.close()

    # 2) Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    # Header row (matches columns)
    headers = ["ID", "Email", "Password", "Full Name", "Username", "DOB", "Bio", "Phone", "Address"]
    ws.append(headers)

    # Data rows
    for r in rows:
        ws.append(list(r))

    # 3) Formatting
    ws.freeze_panes = "A2"

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)

    # 4) Write to memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    debug = os.getenv("FLASK_DEBUG", "0") == "1"
    host = os.getenv("FLASK_HOST", "127.0.0.1")
    port = int(os.getenv("FLASK_PORT", "5000"))
    app.run(debug=debug, host=host, port=port)